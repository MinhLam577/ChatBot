from seleniumCrawl import *
import traceback
import threading
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

root_path = os.path.dirname(os.path.abspath(__file__))
def writeFileTxtResult(fileName, content):
    try:
        with open(fileName, 'a') as f1:
            f1.writelines(content + "\n")
    except Exception:
        traceback.print_exc()
    
def Save_Excel(listpost):
    df = pd.DataFrame(listpost) 
    path = os.path.join(root_path, "DulieuCrawl.xlsx")
    with pd.ExcelWriter(path, mode='w') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    # Load the workbook and select the sheet
    book = load_workbook(path)
    sheet = book['Sheet1']
    # Iterate over the columns
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = 40  # Set the width of all cells to 40
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    # Center align headers
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='center')
    # Top align all other cells and left align content
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', horizontal='left')
    # Save the changes
    book.save(path)

class Threading(threading.Thread):
    def __init__(self, driver, type="fanpage", nameOrID = 103274306376166, numberPost = 100, txt = None, timeout = 60):
        threading.Thread.__init__(self)
        self.driver = driver
        self.type = type #Lọc theo fanpage hoặc group
        self.nameOrId = nameOrID #ID hoặc tên của fanpage hoặc group
        self.numberPost = numberPost #Số lượng bài viết cần lấy trên mỗi fanpage hoặc group
        self.txt = txt #Từ khóa để tìm kiếm bài viết
        self.timeout = timeout #Thời gian kết thức tìm kiếm
    #Bắt đầu tìm kiếm ID của group hoặc fanpage
    def startLocIDGroupOrFanpage(self):
        if(self.type == "fanpage"):
            dataFileID1 = readDataFileITxtID(root_path + "\\" + fileFanpageID)
            dataID1 = get_FangpageID_By_Search(self.driver, self.txt, self.timeout)
            for ID in dataID1:
                if(ID not in dataFileID1):
                    dataFileID1.append(ID)
                    writeFileTxtID(root_path + "\\" + fileFanpageID, ID)
        elif(self.type == "group"):
            dataFileID2 = readDataFileITxtID(root_path + "\\" + fileGroupID)
            dataID2 = get_GroupPublicID_By_Search(self.driver, self.txt, self.timeout)
            for ID in dataID2:
                if(ID not in dataFileID2):
                    dataFileID2.append(ID)
                    writeFileTxtID(root_path + "\\" + fileGroupID, ID)
    #Bắt đầu lấy nội dung bài viết 
    def startGetContentPostBySelenium(self):
        FanpageOrGroupID = self.nameOrId
        cnt = 1
        listpost = []
        try:
            if(self.type == "fanpage"):
                getPostIDFanpage(self.driver, FanpageOrGroupID, self.numberPost)
                listPostFanpageID = readDataFileITxtID(root_path + "\\" + filePostFanpageID)
                with open(root_path + "\\" + filePostFanpageID, 'w') as file:
                    pass
                print("\n\n=================Bắt đầu crawl nội dung================\n\n") 
                for i, ID in enumerate(listPostFanpageID):
                    try:
                        self.driver.get("http://mbasic.facebook.com/"+ID)
                        postContent = getContentFromPostID(self.driver, ID)
                        if(postContent != None):
                            Text = postContent['ContentPost']
                            TimePost = postContent["TimePost"]
                            IDPost = postContent["IDPost"]
                            LinkPost = "https://www.facebook.com/"+IDPost
                            if postContent["LinkImg"] != None:
                                LinkImg = postContent["LinkImg"]
                            else:
                                LinkImg = None
                            print("Crawl thành công bài viết:", cnt, "\n\n","postID:", IDPost, "\nText:",Text, "\nTime:", TimePost, "\npostLink:",LinkPost, "\npostImage:",LinkImg, "\n\n")
                            if(LinkImg != None):
                                download_image(self.driver, LinkImg, IDPost)
                            writeFileTxtResult(root_path + "\\" + filePostFanpageID, LinkPost)
                            listpost.append({
                                "IDPost": IDPost,
                                "TimePost": TimePost,
                                "ContentPost": Text,
                                "LinkPost": LinkPost,
                                "LinkImg": LinkImg
                            })
                            cnt = cnt + 1
                    except AttributeError:
                        raise AttributeError
                    except Exception as e:
                        print(f"Crawl bài viết có ID = {ID} của {self.type} thất bại, lỗi: " + getattr(e, 'message', repr(e)))
                if len(listpost) > 0:
                    Save_Excel(listpost)
            elif(self.type == "group"):
                getPostsIDGroup(self.driver, FanpageOrGroupID, self.numberPost)
                listPostGroupID = readDataFileITxtID(root_path + "\\" +filePostGroupID)
                with open(root_path + "\\" + filePostGroupID, 'w') as file:
                    pass
                print("\n\n=================Bắt đầu crawl nội dung================\n\n")
                for i, ID in enumerate(listPostGroupID):
                    try:
                        self.driver.get("https://mbasic.facebook.com/"+ID)
                        postContent = getContentFromPostID(self.driver, ID)
                        Text = postContent['ContentPost']
                        TimePost = postContent["TimePost"]
                        IDPost = postContent["IDPost"]
                        LinkPost = "https://www.facebook.com/"+IDPost
                        if postContent["LinkImg"] != None:
                            LinkImg = postContent["LinkImg"]
                        else:
                            LinkImg = None
                        print("Crawl thành công bài viết:",cnt, "\n\n","postID:", IDPost, "\nText:",Text, "\nTime:", TimePost,
                                "\npostLink:",LinkPost, "\npostImage:",LinkImg, "\n\n")
                        listpost.append({
                            "IDPost": IDPost,
                            "TimePost": TimePost,
                            "ContentPost": Text,
                            "LinkPost": LinkPost,
                            "LinkImg": LinkImg
                        })
                        writeFileTxtResult(root_path + "\\" + filePostGroupID, LinkPost)
                        if(LinkImg != None):
                            download_image(self.driver, LinkImg, IDPost)
                        cnt = cnt + 1
                    except AttributeError:
                        raise AttributeError
                    except Exception as e:
                        print(f"Crawl bài viết có ID = {ID} của {self.type} thất bại, lỗi: " + getattr(e, 'message', repr(e)))
                if len(listpost) > 0:
                    Save_Excel(listpost)
        except PermissionError:
            print("Vui lòng đóng file excel trước khi chạy chương trình")
        except NoSuchElementException:
            print("Không tìm thấy Fanpages hoặc group, vui lòng kiếm tra lại ID hoặc tên đã nhập")
        except AttributeError:
            if len(listpost) > 0:
                Save_Excel(listpost)
            print("Đã bị ban")
    def closeDriverProfile(self):
        self.driver.close()
    def run(self):
        if(self.txt != None):
            self.startLocIDGroupOrFanpage()
        elif(self.txt == None):
            self.startGetContentPostBySelenium() 
            
def LocIDFanpageAndGroupPost(*,driver1, driver2, cookie1, cookie2, type1 = 'fanpage', type2 = 'group', txt, timeout = 60):
    try:
        driver1 = initDriverProfile("--headless=new")
        driver2 = initDriverProfile("--headless=new")
        loginFacebookByCookie(driver1, cookie1)
        loginFacebookByCookie(driver2, cookie2)
        thread1 = Threading(driver1, cookie=cookie1, type=type1, txt=txt, timeout=timeout)
        thread2 = Threading(driver2, cookie=cookie2, type=type2, txt=txt, timeout=timeout)
        thread1.start()
        thread2.start()
        thread1.join()
        thread2.join()
    except Exception:
        traceback.print_exc()
    finally:
        thread1.closeDriverProfile()
        thread2.closeDriverProfile()

def getContentPostFanpageOrGroupBySelenium(*,driver, type = 'fanpage', NameOrID, numberPost = 100):
    try:
        thread = Threading(driver, type=type, numberPost=numberPost, nameOrID=NameOrID)
        thread.start()
        thread.join()
    except Exception as e:
        print("Lấy nội dung bài viết thất bại, lỗi: " + getattr(e, 'message', repr(e)))
        
def clearData(path):
    with open(os.path.join(path, filePostFanpageID), 'w') as file:
        pass
    with open(os.path.join(path, filePostGroupID), 'w') as file:
        pass
    path = os.path.join(path, "DulieuCrawl.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    # Xóa tất cả các hàng
    ws.delete_rows(1, ws.max_row)
    # Save the changes
    wb.save(path)

def startGetContentPostBySelenium(driver, type: str = "fanpage", NameOrID: str = None, numberPost: int = 10):
    try:
        path = root_path
        clearData(path)
        getContentPostFanpageOrGroupBySelenium(driver=driver, type=type, numberPost=numberPost, NameOrID=NameOrID)
        if(type == "fanpage"):
            print(f"Đã lấy xong bài viết của {type} có", "NameOrID là:", NameOrID, "vui lòng kiểm tra file " + filePostFanpageID + " để xem các link bài viết đã lấy, để biết thêm chi tiết về bài viết vui lòng kiểm tra file DulieuCrawl.xlsx")
        elif(type == "group"):
            print(f"Đã lấy xong bài viết của {type} có", "NameOrID là:", NameOrID, "vui lòng kiểm tra file " + filePostGroupID + " để xem các link bài viết đã lấy, để biết thêm chi tiết về bài viết vui lòng kiểm tra file DulieuCrawl.xlsx")
    except PermissionError:
        print("Vui lòng đóng file DulieuCrawl.xlsx trước khi chạy chương trình")
    except Exception:
        traceback.print_exc()        

try:
    driver = initDriverProfile()
    cookie = getCookieFromFile("cookies.txt")
    loginFacebookByCookie(driver, cookie)
    if(driver.title.__contains__("log in") or driver.title == ""):
        raise IndexError
    while True:
        try:
            Check_type = False
            print("================================================")
            print("Các chức năng chính của chương trình: ")
            print("0. Crawl bài viết từ Fanpages")
            print("1. Crawl bài viết từ group")
            print("2. Lọc ID Fanpages theo từ khóa tìm kiếm")
            print("3. Lọc ID Group public theo từ khóa tìm kiếm")
            print("4. Lọc ID Group đã gia nhập(join) theo từ khóa tìm kiếm")
            print("5. Thoát chương trình")
            print("================================================")
            type = None
            while not Check_type:
                type = input("Nhập vào lựa chọn: ")
                Check_type = bool(re.match('^[0-5]$', type))
                if not Check_type:
                    print("Vui lòng nhập các số nguyên từ 0 đến 5")
            if(type == "0"):
                NameOrID = input("Nhập vào ID hoặc tên của fanpage: ")
                Check_type = False
                Count = None
                while not Check_type:
                    Count = input("Nhập vào số lượng bài viết cần crawl: ")
                    Check_type = bool(re.match('^\d+$', Count))
                    if not Check_type:
                        print("Số lượng bài viết phải là số nguyên")
                    else:
                        if(int(Count) <= 0):
                            print("Số lượng bài viết phải lớn hơn 0")
                startGetContentPostBySelenium(driver, 'fanpage', NameOrID, int(Count))
            elif type == "1":
                NameOrID = input("Nhập vào ID hoặc tên của group: ")
                Check_cnt = False
                Count = None
                while not Check_cnt:
                    Count = input("Nhập vào số lượng bài viết cần crawl: ")
                    Check_cnt = bool(re.match('^\d+$', Count))
                    if not Check_cnt:
                        print("Số lượng bài viết phải là số nguyên")
                    else:
                        if(int(Count) <= 0):
                            print("Số lượng bài viết phải lớn hơn 0")
                startGetContentPostBySelenium(driver, 'group', NameOrID, int(Count))
            elif type == "2":
                Key = input("Nhập vào từ khóa fanpage cần tìm: ")
                Check_cnt = False
                Count = None
                while not Check_cnt:
                    Count = input("Nhập vào khoảng thời gian tìm kiếm(giây) cố định là 15 giây: ")
                    Check_cnt = bool(re.match('^\d+$', Count))
                    if not Check_cnt:
                        print("Thời gian phải là số nguyên")
                    else:
                        if(int(Count) <= 0):
                            print("Thời gian phải lớn hơn 0")
                res = get_FangpageID_By_Search(driver, Key, Count)
                print(f"Danh sách các ID Fanpages đã lọc được lưu vào trong file {fileFanpageID}: \n{res}\n\n")
                store_res = readDataFileITxtID(root_path + "\\" + fileFanpageID)
                for i in res:
                    if i not in store_res:
                        writeFileTxtID(root_path + "\\" + fileFanpageID, i)
            elif type == "3":
                Key = input("Nhập vào từ khóa group cần tìm: ")
                Check_cnt = False
                Count = None
                while not Check_cnt:
                    Count = input("Nhập vào khoảng thời gian tìm kiếm(giây) cố định là 15 giây: ")
                    Check_cnt = bool(re.match('^\d+$', Count))
                    if not Check_cnt:
                        print("Thời gian phải là số nguyên")
                    else:
                        if(int(Count) <= 0):
                            print("Thời gian phải lớn hơn 0")
                res = get_GroupPublicID_By_Search(driver, Key, Count)
                print(f"Danh sách các ID group đã lọc được lưu vào file {fileGroupID}: \n{res}\n\n")
                store_res = readDataFileITxtID(root_path + "\\" + fileGroupID)
                for i in res:
                    if i not in store_res:
                        writeFileTxtID(root_path + "\\" + fileGroupID, i)
            elif type == "4":
                Key = input("Nhập vào từ khóa group cần tìm: ")
                Check_cnt = False
                Count = None
                while not Check_cnt:
                    Count = input("Nhập vào khoảng thời gian tìm kiếm(giây) cố định là 15 giây: ")
                    Check_cnt = bool(re.match('^\d+$', Count))
                    if not Check_cnt:
                        print("Thời gian phải là số nguyên")
                    else:
                        if(int(Count) <= 0):
                            print("Thời gian phải lớn hơn 0")
                res = get_JoinedGroupID_By_Search(driver, Key, Count)
                print(f"Danh sách các ID group đã lọc được lưu vào file {fileGroupIDJoin}: \n{res}\n\n")
                store_res = readDataFileITxtID(root_path + "\\" + fileGroupIDJoin)
                for i in res:
                    if i not in store_res:
                        writeFileTxtID(root_path + "\\" + fileGroupIDJoin, i)
            elif type == "5":
                break
            input("Nhấn Enter để tiếp tục...")
            os.system('cls' if os.name == 'nt' else 'clear')
        except NoSuchElementException:
            print("Không tìm thấy bài viết cần tìm")
        except Exception as e:
            print("Lỗi chương trình: ", getattr(e, 'message', repr(e)))
except IndexError:
    print("Cookie không hợp lệ")
    input("Nhấn Enter để tiếp tục...")
os.system('cls' if os.name == 'nt' else 'clear')